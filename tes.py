"""
Script debug: cek isi file ANALISIS NON SETTING yang diupload
Jalankan: python cek_previous.py
"""

import json
import os
from openpyxl import load_workbook
from datetime import datetime

# ── 1. Baca INVENTRA.json untuk dapat path file ──────────────────────────────
if not os.path.exists("INVENTRA.json"):
    print("[ERROR] INVENTRA.json tidak ditemukan. Jalankan script ini di folder yang sama dengan main.py")
    exit()

with open("INVENTRA.json", "r") as f:
    cfg = json.load(f)

ns_mapping = cfg.get("data_mappings", {}).get("ANALISIS NON SETTING", {})
ns_file_id = ns_mapping.get("source_file_id")
print(f"source_file_id untuk ANALISIS NON SETTING : {ns_file_id!r}")

uploaded_files = cfg.get("uploaded_files", {})
file_info = uploaded_files.get(ns_file_id, {})
file_path = file_info.get("path")
print(f"Path file                                  : {file_path!r}")

if not file_path:
    print("[ERROR] Path file kosong — pastikan mapping ANALISIS NON SETTING sudah dikonfigurasi.")
    exit()

if not os.path.exists(file_path):
    print(f"[ERROR] File tidak ditemukan di path: {file_path}")
    exit()

# ── 2. Buka workbook & list semua sheet ──────────────────────────────────────
print("\n── Sheet list ──────────────────────────────────────────────────────────")
wb = load_workbook(file_path, read_only=True, data_only=True)
for i, name in enumerate(wb.sheetnames):
    print(f"  [{i}] {name!r}")

# ── 3. Cari sheet COVER ──────────────────────────────────────────────────────
cover_name = next((s for s in wb.sheetnames if "COVER" in s.upper()), None)
print(f"\nSheet COVER ditemukan : {cover_name!r}")

if not cover_name:
    print("[ERROR] Tidak ada sheet yang mengandung kata 'COVER'.")
    wb.close()
    exit()

# ── 4. Baca area sekitar C7 ──────────────────────────────────────────────────
ws = wb[cover_name]
print("\n── Isi sel area B5:E10 (untuk cek posisi tanggal) ──────────────────────")
for row in ws.iter_rows(min_row=5, max_row=10, min_col=2, max_col=5, values_only=True):
    for ci, val in enumerate(row, start=2):
        col_letter = chr(ord('A') + ci - 1)
        row_num = 5 + list(ws.iter_rows(min_row=5, max_row=10, values_only=True)).index(
            ws.iter_rows(min_row=5, max_row=10, min_col=2, max_col=5, values_only=True).__next__()
            if False else row
        ) if False else None
    print(f"  {row}")

# Cara lebih simpel: langsung print baris 5-10
print("\n── Baris 5–10, kolom A–F ───────────────────────────────────────────────")
for r in ws.iter_rows(min_row=5, max_row=10, min_col=1, max_col=6, values_only=True):
    print(f"  {r}")

# ── 5. Cek spesifik C7 ───────────────────────────────────────────────────────
cell_c7 = ws["C7"].value
print(f"\n── Nilai cell C7 ───────────────────────────────────────────────────────")
print(f"  value : {cell_c7!r}")
print(f"  type  : {type(cell_c7)}")

if isinstance(cell_c7, datetime):
    print(f"  ✅ Sudah datetime → {cell_c7.strftime('%d %B %Y, %H:%M')}")
elif cell_c7:
    import pandas as pd
    parsed = pd.to_datetime(str(cell_c7), dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        print(f"  ⚠️  String tapi gagal di-parse sebagai tanggal: {cell_c7!r}")
        print(f"     Coba format lain atau cek apakah sel ini merged/formula.")
    else:
        print(f"  ✅ String berhasil di-parse → {parsed.strftime('%d %B %Y, %H:%M')}")
else:
    print("  ❌ Cell C7 KOSONG — tanggal tidak bisa dibaca.")
    print("     Kemungkinan: sel ini merged, atau tanggal ada di sel lain.")

wb.close()
print("\n── Selesai ─────────────────────────────────────────────────────────────")