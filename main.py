import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import data
import pandas as pd
from datetime import datetime
import json
import os
from PIL import Image,ImageTk
import sys
from tutorial import TutorialScreen
from drp_app import DRPApp
from proses_GSheet import WEB_APP_URL
import threading

try:
    import pywinstyles
    PYWINSTYLES_AVAILABLE = True
except ImportError:
    PYWINSTYLES_AVAILABLE = False

def apply_titlebar_color(window, color="#7F1D1D"):
    """Terapkan warna title bar sesuai tema INVENTRA (Windows only)"""
    if not PYWINSTYLES_AVAILABLE:
        return
    try:
        pywinstyles.change_header_color(window, color)
    except Exception:
        pass

def resource_path(relative_path):
    """Return path yang benar saat jadi EXE atau dijalankan di Python biasa"""
    try:
        # Saat PyInstaller bikin exe, file diekstrak ke folder _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class MappingDialog:
    def __init__(self, parent, colors, uploaded_files, callback, overlay=None):
        self.parent = parent
        self.colors = colors
        self.uploaded_files = uploaded_files
        self.callback = callback
        self.overlay = overlay
        
        # Load existing settings
        self.settings = self.load_settings()
        
        # Create dialog
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Data Mapping Configuration")
        self.dialog.overrideredirect(True)
        self.dialog.configure(bg="#f5f6fa")

        # posisi & fokus
        self.dialog.transient(parent)
        self.dialog.lift()
        self.dialog.grab_set()
        self.dialog.focus_force()
        self.dialog.update_idletasks()
        self.dialog.bind("<Escape>", lambda e: self.close_dialog())

        # ===== FIX ALT+TAB FOCUS LOST =====
        self.dialog.bind("<FocusIn>", self._force_focus)
        self.parent.bind("<FocusIn>", self._force_focus)

        
        # Get screen dimensions
        screen_width = self.dialog.winfo_screenwidth()
        screen_height = self.dialog.winfo_screenheight()
        
        # Set dialog size (80% of screen)
        dialog_width = int(screen_width * 0.8)
        dialog_height = int(screen_height * 0.85)
        
        self.dialog.configure(bg=colors['bg'])
        self.dialog.transient(parent)
        
        # Center window
        x = (screen_width - dialog_width) // 2
        y = (screen_height - dialog_height) // 2
        self.dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        # Prevent resize
        self.dialog.resizable(False, False)
        
        self.create_ui()

    def _force_focus(self, event=None):
        # 🔥 Cek dulu apakah dialog masih ada
        if not hasattr(self, "dialog") or not self.dialog.winfo_exists():
            return

        widget = self.dialog.focus_get()

        # Jangan ganggu saat user sedang mengetik di Entry/Text
        if isinstance(widget, (tk.Entry, tk.Text)):
            return

        self.dialog.focus_force()


        
    def load_settings(self):
        """Load settings dari JSON"""
        default_settings = {
            "data_mappings": {
                "PLJM01": {"source_file_id": None, "sheet_name": None, "columns": {}},
                "PLJM08": {"source_file_id": None, "sheet_name": None, "columns": {}},
                "SRD": {"source_file_id": None, "sheet_name": None, "columns": {}},
                "SLN": {"source_file_id": None, "sheet_name": None, "columns": {}},
                "IR": {"source_file_id": None, "sheet_name": None, "columns": {}},
                "PO": {"source_file_id": None, "sheet_name": None, "columns": {}},
                "LEVERING": {"source_file_id": None, "sheet_name": None, "columns": {}},
                "ANALISIS SETTING": {"source_file_id": None, "sheet_name": None, "columns": {}},
                "ANALISIS NON SETTING": {"source_file_id": None, "sheet_name": None, "columns": {}},
            }
        }
        
        try:
            if os.path.exists("INVENTRA.json"):
                with open("INVENTRA.json", "r") as f:
                    loaded = json.load(f)
                    if "data_mappings" in loaded:
                        for key in default_settings["data_mappings"]:
                            if key in loaded["data_mappings"]:
                                default_settings["data_mappings"][key].update(loaded["data_mappings"][key])
        except Exception as e:
            print(f"Error loading settings: {e}")
        
        return default_settings
    
    def save_settings(self):
        """Save settings ke JSON"""
        try:
            # Load existing settings
            existing = {}
            if os.path.exists("INVENTRA.json"):
                with open("INVENTRA.json", "r") as f:
                    existing = json.load(f)
            
            # Update data_mappings
            existing["data_mappings"] = self.settings["data_mappings"]
            
            # Save
            with open("INVENTRA.json", "w") as f:
                json.dump(existing, f, indent=4)
        except Exception as e:
            print(f"Error saving settings: {e}")
    
    def create_ui(self):
        # Header
        header = tk.Frame(self.dialog, bg=self.colors['primary'], height=70)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        header_content = tk.Frame(header, bg=self.colors['primary'])
        header_content.pack(expand=True, fill="both", padx=30)
        
        tk.Label(
            header_content,
            text="Configure Data Mapping",
            font=("Segoe UI", 18, "bold"),
            bg=self.colors['primary'],
            fg="white"
        ).pack(side="left")
        
        tk.Label(
            header_content,
            text="Map your Excel files to data sources",
            font=("Segoe UI", 10),
            bg=self.colors['primary'],
            fg="#fecaca"
        ).pack(side="right")
        
        # Main content area
        content_frame = tk.Frame(self.dialog, bg=self.colors['bg'])
        content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Create notebook for tabs
        style = ttk.Style()
        style.theme_use('default')
        style.configure('TNotebook', background=self.colors['bg'], borderwidth=0)
        style.configure('TNotebook.Tab', 
                       padding=[20, 10], 
                       font=('Segoe UI', 10, 'bold'))
        style.map('TNotebook.Tab',
                 background=[('selected', self.colors['primary'])],
                 foreground=[('selected', 'white'), ('!selected', self.colors['text'])])
        
        self.notebook = ttk.Notebook(content_frame)
        self.notebook.pack(fill="both", expand=True)
        
        # Create tabs for each data type
        data_types = ["PLJM01", "PLJM08", "SRD", "SLN", "IR", "PO", "LEVERING", "ANALISIS SETTING", "ANALISIS NON SETTING"]
        self.widgets = {}
        
        for data_type in data_types:
            tab_frame = tk.Frame(self.notebook, bg=self.colors['card_bg'])
            self.notebook.add(tab_frame, text=f"  {data_type}  ")
            self.create_mapping_tab(tab_frame, data_type)
        
        # Buttons frame
        button_frame = tk.Frame(self.dialog, bg=self.colors['bg'], height=70)
        button_frame.pack(fill="x", padx=20, pady=(0, 20))
        button_frame.pack_propagate(False)
        
        button_container = tk.Frame(button_frame, bg=self.colors['bg'])
        button_container.pack(expand=True)
        
        cancel_btn = tk.Button(
            button_container,
            text="Cancel",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors['text_light'],
            fg="white",
            relief="flat",
            cursor="hand2",
            command=self.close_dialog,
            padx=30,
            pady=12
        )
        cancel_btn.pack(side="left", padx=5)
        
        cancel_btn.bind("<Enter>", lambda e: cancel_btn.config(bg="#4b5563"))
        cancel_btn.bind("<Leave>", lambda e: cancel_btn.config(bg=self.colors['text_light']))
        
        lanjut_btn = tk.Button(
            button_container,
            text="✓ Lanjut",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors['primary'],
            fg="white",
            relief="flat",
            cursor="hand2",
            command=self.on_lanjut,
            padx=30,
            pady=12
        )
        lanjut_btn.pack(side="left", padx=5)
        
        lanjut_btn.bind("<Enter>", lambda e: lanjut_btn.config(bg=self.colors['primary_dark']))
        lanjut_btn.bind("<Leave>", lambda e: lanjut_btn.config(bg=self.colors['primary']))
    
    def create_mapping_tab(self, parent, data_type):
        """Create mapping tab for one data type"""
        # Create scrollable area
        canvas = tk.Canvas(parent, bg=self.colors['card_bg'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        
        scrollable_frame = tk.Frame(canvas, bg=self.colors['card_bg'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Bind mousewheel
        def _on_mousewheel(event):
            if canvas.winfo_exists():
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        def _bind(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind(event):
            canvas.unbind_all("<MouseWheel>")

        canvas.bind("<Enter>", _bind)
        canvas.bind("<Leave>", _unbind)

        # Main content
        content = tk.Frame(scrollable_frame, bg=self.colors['card_bg'])
        content.pack(fill="both", expand=True, padx=40, pady=30)
        
        # Info header
        info_frame = tk.Frame(content, bg=self.colors['primary_light'], relief="flat", bd=0)
        info_frame.pack(fill="x", pady=(0, 25))
        
        tk.Label(
            info_frame,
            text=f"📋 Configure {data_type} Data Source",
            font=("Segoe UI", 14, "bold"),
            bg=self.colors['primary_light'],
            fg=self.colors['text'],
            anchor="w"
        ).pack(fill="x", padx=20, pady=15)
        
        # Source file and sheet selection
        selection_frame = tk.Frame(content, bg="white", relief="solid", bd=1)
        selection_frame.pack(fill="x", pady=(0, 20))
        
        selection_content = tk.Frame(selection_frame, bg="white")
        selection_content.pack(fill="x", padx=25, pady=20)
        
        # Source File Row
        tk.Label(
            selection_content,
            text="Source File",
            font=("Segoe UI", 11, "bold"),
            bg="white",
            fg=self.colors['text'],
            anchor="w"
        ).pack(fill="x", pady=(0, 8))
        
        uploaded_files_list = [f"{self.uploaded_files[fid]['filename']}" 
                               for fid in self.uploaded_files.keys()]
        self.filename_to_id = {
            self.uploaded_files[fid]['filename']: fid
            for fid in self.uploaded_files
        }
        
        current_mapping = self.settings["data_mappings"].get(data_type, {})
        # source_file_var = tk.StringVar(value=current_mapping.get("source_file_id", ""))
        saved_file_id = current_mapping.get("source_file_id", "")
        saved_filename = ""

        if saved_file_id in self.uploaded_files:
            saved_filename = self.uploaded_files[saved_file_id]['filename']

        source_file_var = tk.StringVar(value=saved_filename)
        
        source_file_combo = ttk.Combobox(
            selection_content,
            textvariable=source_file_var,
            state="readonly",
            values=uploaded_files_list,
            font=("Segoe UI", 10),
            height=8
        )
        source_file_combo.pack(fill="x", pady=(0, 15))
        
        # Sheet Name Row
        tk.Label(
            selection_content,
            text="Sheet Name",
            font=("Segoe UI", 11, "bold"),
            bg="white",
            fg=self.colors['text'],
            anchor="w"
        ).pack(fill="x", pady=(0, 8))
        
        sheet_name_var = tk.StringVar(value=current_mapping.get("sheet_name", ""))
        
        sheet_name_combo = ttk.Combobox(
            selection_content,
            textvariable=sheet_name_var,
            state="readonly",
            values=[],
            font=("Segoe UI", 10),
            height=8
        )
        sheet_name_combo.pack(fill="x")

        header_vars = {}
        
        def on_file_selected(event=None):
            selected = source_file_var.get()
            if selected:
                file_id = self.filename_to_id.get(selected)
                print(selected)
                print("file id:", file_id)
                if file_id in self.uploaded_files:
                    sheets = list(self.uploaded_files[file_id]['sheets'].keys())
                    sheet_name_combo['values'] = sheets
                    if sheets and not sheet_name_var.get():
                        sheet_name_combo.current(0)
            
            update_header_dropdowns()

        def update_header_dropdowns():
            selected_file = source_file_var.get()
            selected_sheet = sheet_name_var.get()

            if not selected_file or not selected_sheet:
                return

            file_id = self.filename_to_id.get(selected_file)
            if file_id not in self.uploaded_files:
                return

            df = self.uploaded_files[file_id]['sheets'].get(selected_sheet)
            if df is None:
                return

            columns = list(df.columns)

            for field_key, obj in header_vars.items():
                obj["widget"]["values"] = columns

                # Reset jika value lama tidak ada di kolom baru
                if obj["var"].get() not in columns:
                    obj["var"].set("")

        
        source_file_combo.bind("<<ComboboxSelected>>", on_file_selected)
        sheet_name_combo.bind("<<ComboboxSelected>>", lambda e: update_header_dropdowns())
        
        # Initialize if already has value
        if source_file_var.get():
            on_file_selected()
            if sheet_name_var.get():
                sheet_name_combo.set(sheet_name_var.get())
        
        # Column headers section
        tk.Label(
            content,
            text="Column Headers Configuration",
            font=("Segoe UI", 12, "bold"),
            bg=self.colors['card_bg'],
            fg=self.colors['text'],
            anchor="w"
        ).pack(fill="x", pady=(10, 10))
        
        tk.Label(
            content,
            text="Specify the exact column names in your Excel sheet (optional - leave default if matches)",
            font=("Segoe UI", 9),
            bg=self.colors['card_bg'],
            fg=self.colors['text_light'],
            anchor="w"
        ).pack(fill="x", pady=(0, 15))
        
        headers_frame = tk.Frame(content, bg="white", relief="solid", bd=1)
        headers_frame.pack(fill="both", expand=True)
        
        headers_content = tk.Frame(headers_frame, bg="white")
        headers_content.pack(fill="both", expand=True, padx=25, pady=20)
        
        # Get column config
        column_configs = self.get_column_config(data_type)
        current_columns = current_mapping.get("columns", {})
        
        # Create grid layout for headers
        for idx, (field_key, field_label, default_value) in enumerate(column_configs):
            row_frame = tk.Frame(headers_content, bg="white")
            row_frame.pack(fill="x", pady=6)
            
            label_frame = tk.Frame(row_frame, bg="white", width=200)
            label_frame.pack(side="left", fill="y")
            label_frame.pack_propagate(False)
            
            tk.Label(
                label_frame,
                text=field_label + ":",
                font=("Segoe UI", 10),
                bg="white",
                fg=self.colors['text'],
                anchor="w"
            ).pack(side="left", padx=(0, 10))
            
            var = tk.StringVar(value=current_columns.get(field_key, default_value))
            header_vars[field_key] = var
            
            combo = ttk.Combobox(
                row_frame,
                textvariable=var,
                state="readonly",
                font=("Segoe UI", 10),
                values=[]   # nanti kita isi
            )
            combo.pack(side="left", fill="x", expand=True)

            header_vars[field_key] = {
                "var": var,
                "widget": combo
            }

        # Store widgets
        self.widgets[data_type] = {
            'source_file_var': source_file_var,
            'sheet_name_var': sheet_name_var,
            'header_vars': header_vars
        }
    
    def get_column_config(self, data_type):
        """Get column configuration for each data type"""
        configs = {
            "PLJM01": [
                ("stock_code", "Stock Code", "Stock Code"),
                ("rop", "ROP", "ROP"),
                ("roq", "ROQ", "ROQ"),
            ],
            "PLJM08": [
                ("distric", "Distric", "Distric"),
                ("stock_code", "Stock Code", "Stock Code"),
                ("item_name", "Item Name", "Item Name"),
                ("soh_akhir", "SOH Akhir", "SOH Akhir"),
                ("exp", "EXP", "Exp")
            ],
            "SRD": [
                ("stock_code", "Stock Code", "STOCK_CODE"),
                ("qty_rcv_uop", "Qty Received", "QTY_RCV_UOP"),
                ("creation_date", "Creation Date", "CREATION_DATE")
            ],
            "SLN": [
                ("stock_code", "Stock Code", "Stock Code"),
                ("qty_req", "Qty Required", "Qty Req"),
                ("qty_issued", "Qty Issued", "QTY_ISSUED"),
                ("last_acq_date", "Last Acq Date", "Last Acq Date")
            ],
            "IR": [
                ("stock_code", "Stock Code", "Stock Code"),
                ("qty_req", "Qty Required", "Qty Req"),
                ("qty_issued", "Qty Issued", "QTY_ISSUED"),
                ("req_by_date", "Req By Date", "Req By Date")
            ],
            "PO": [
                ("stock_code", "Stock Code", "STOCK_CODE"),
                ("supplier_name", "Supplier Name", "SUPPLIER_NAME"),
                ("order_date", "Order Date", "ORDER_DATE"),
                ("receipt_status", "Receipt Status", "RECEIPT_STATUS"),
                ("qty_rcv_dir", "Qty Rcv Dir", "QTY_RCV_DIR"),
                ("curr_qty", "Curr Qty", "CURR_QTY")
            ],
            "LEVERING": [
                ("stock_code", "Stock Code", "STOCK_CODE"),
                ("receipt_status", "Receipt Status", "RECEIPT_STATUS"),
                ("curr_qty_p", "Current Qty", "CURR_QTY_P"),
                ("due_site_date", "Due Site Date", "DUE_SITE_DATE")
            ],
            "ANALISIS SETTING": [
                ("supplier_name", "Supplier Name", "Supplier Name"),
                ("stock_code", "Stock Code", "STOCK_CODE"),
                ("item_name", "Item Name", "Item Name"),
                ("analisis", "Analisis", "Keterangan")
            ],
            "ANALISIS NON SETTING": [
                ("supplier_name", "Supplier Name", "Supplier Name"),
                ("stock_code", "Stock Code", "STOCK_CODE"),
                ("item_name", "Item Name", "Item Name"),
                ("analisis", "Analisis", "Keterangan")
            ]
        }
        
        return configs.get(data_type, [])
    
    def on_lanjut(self):
        """Validate and save mappings"""

        mappings = {}

        for data_type, widgets in self.widgets.items():
            source_file = widgets['source_file_var'].get()
            sheet_name = widgets['sheet_name_var'].get()

            if not source_file or not sheet_name:
                messagebox.showwarning(
                    "Incomplete Mapping",
                    f"Please select source file and sheet for {data_type}"
                )
                return

            file_id = self.filename_to_id.get(source_file)

            headers = {}
            for field_key, obj in widgets['header_vars'].items():
                headers[field_key] = obj["var"].get()

            mappings[data_type] = {
                "source_file_id": file_id,
                "sheet_name": sheet_name,
                "columns": headers
            }

        # update settings
        self.settings["data_mappings"] = mappings

        # save JSON dulu
        self.save_settings()

        # 🔥 JALANKAN PROCESS DULU
        self.callback(mappings)

        # 🔥 BARU TUTUP DIALOG
        self.close_dialog()


    def close_dialog(self):
        if self.overlay and self.overlay.winfo_exists():
            self.overlay.destroy()

        if self.dialog and self.dialog.winfo_exists():
            self.dialog.destroy()



class ModernRedINVENTRAManager:
    def __init__(self, root):
        self.root = root
        try:
            self.root.iconbitmap(resource_path("logo_app.ico"))
        except Exception as e:
            print("Icon tidak ditemukan:", e)
        try:
            self.root.iconphoto(True, tk.PhotoImage(file=resource_path("logo_app.png")))
        except:
            pass
        self.root.title("INVENTRA")
        self.root.attributes("-fullscreen", True)
        self.root.configure(bg="#fafafa")
        self.root.bind("<Escape>", lambda e: self.root.quit())
        
        # Data store untuk file yang diupload
        self.uploaded_files = {}
        self.file_counter = 0
        
        # Data store untuk data yang sudah dimapping
        self.data_store = {
            "PLJM01": None,
            "PLJM08": None,
            "SRD": None,
            "SLN": None,
            "IR": None,
            "PO": None,
            "LEVERING": None,
            "ANALISIS SETTING": None,
            "ANALISIS NON SETTING": None,
        }
        
        # Current selected view
        self.current_view = None
        
        # Modern Red Color Scheme
        self.colors = {
            'primary': '#dc2626',
            'primary_light': '#fee2e2',
            'primary_dark': '#991b1b',
            'accent': '#ef4444',
            'success': '#10b981',
            'warning': '#f59e0b',
            'bg': '#fafafa',
            'card_bg': '#ffffff',
            'text': '#1f2937',
            'text_light': '#6b7280',
            'border': '#e5e7eb',
            'shadow': '#00000010'
        }

        # ── Custom Title Bar (permanen di atas, setelah colors didefinisikan) ──
        self._build_custom_titlebar()
        
        # Load settings dan uploaded files dari JSON
        # self.load_from_json()
        
        # Show splash screen first
        self.show_splash_screen()
        
    def _build_custom_titlebar(self):
        """Custom title bar merah permanen di bagian paling atas window"""
        c = self.colors
        BAR_H = 32

        bar = tk.Frame(self.root, bg=c['primary_dark'], height=BAR_H)
        bar.place(x=0, y=0, relwidth=1, height=BAR_H)
        bar.lift()
        bar.lower()   # taruh di bawah konten lain agar tidak menghalangi
        self._titlebar = bar

        # ── Ikon + Nama Aplikasi ──
        left = tk.Frame(bar, bg=c['primary_dark'])
        left.pack(side="left", padx=(10, 0), fill="y")

        try:
            ico = Image.open(resource_path("logo_app.png"))
            ico.thumbnail((20, 20), Image.LANCZOS)
            self._tb_ico = ImageTk.PhotoImage(ico)
            tk.Label(left, image=self._tb_ico,
                     bg=c['primary_dark']).pack(side="left", padx=(0, 6))
        except Exception:
            pass

        tk.Label(
            left,
            text="INVENTRA",
            font=("Segoe UI", 10, "bold"),
            bg=c['primary_dark'],
            fg="white"
        ).pack(side="left")

        # ── Tombol kanan: Minimize · Maximize · Close ──
        right = tk.Frame(bar, bg=c['primary_dark'])
        right.pack(side="right", fill="y")

        def _btn(parent, text, cmd, hover_bg):
            b = tk.Label(
                parent, text=text,
                font=("Segoe UI", 11),
                bg=c['primary_dark'], fg="white",
                padx=14, pady=4, cursor="hand2"
            )
            b.pack(side="left")
            b.bind("<Enter>",          lambda e: b.config(bg=hover_bg))
            b.bind("<Leave>",          lambda e: b.config(bg=c['primary_dark']))
            b.bind("<ButtonRelease-1>", lambda e: cmd())
            return b

        def _minimize():
            self.root.attributes("-fullscreen", False)
            self.root.iconify()

        _btn(right, "─",  _minimize,       c['primary'])
        _btn(right, "✕",  self.root.quit,  "#B91C1C")

        # Saat window dipanggil lagi dari taskbar → fullscreen kembali
        self.root.bind("<Map>", lambda e: self.root.after(
            100, lambda: self.root.attributes("-fullscreen", True)
        ))

        # Drag window saat bukan fullscreen
        bar.bind("<ButtonPress-1>",   self._tb_start_drag)
        bar.bind("<B1-Motion>",       self._tb_drag)
        left.bind("<ButtonPress-1>",  self._tb_start_drag)
        left.bind("<B1-Motion>",      self._tb_drag)

        # Pastikan bar selalu di atas saat pertama dibuat
        self.root.update_idletasks()
        bar.lift()

    def _tb_start_drag(self, event):
        self._drag_x = event.x_root
        self._drag_y = event.y_root

    def _tb_drag(self, event):
        if self.root.attributes("-fullscreen"):
            return
        dx = event.x_root - self._drag_x
        dy = event.y_root - self._drag_y
        x  = self.root.winfo_x() + dx
        y  = self.root.winfo_y() + dy
        self.root.geometry(f"+{x}+{y}")
        self._drag_x = event.x_root
        self._drag_y = event.y_root

    def _lift_titlebar(self):
        """Angkat title bar ke atas setelah layar baru ditampilkan"""
        try:
            if hasattr(self, "_titlebar") and self._titlebar.winfo_exists():
                self._titlebar.lift()
        except Exception:
            pass

    def load_from_json(self):
        """Load settings dan uploaded files dari JSON"""
        try:
            if os.path.exists("INVENTRA.json"):
                with open("INVENTRA.json", "r") as f:
                    settings = json.load(f)
                    
                    # Load uploaded files info
                    if "uploaded_files" in settings:
                        saved_files = settings["uploaded_files"]
                        for file_id, file_info in saved_files.items():
                            file_path = file_info.get("path")
                            if file_path and os.path.exists(file_path):
                                try:
                                    # Reload Excel file
                                    excel_file = pd.ExcelFile(file_path)
                                    sheets_data = {}
                                    for sheet_name in excel_file.sheet_names:
                                        df = self.read_excel_auto_header(file_path, sheet_name)
                                        sheets_data[sheet_name] = df
                                    
                                    self.uploaded_files[file_id] = {
                                        'path': file_path,
                                        'sheets': sheets_data,
                                        'size': file_info.get('size', 0),
                                        'filename': file_info.get('filename', '')
                                    }
                                    
                                    # Update file counter
                                    if file_id.startswith("file_"):
                                        num = int(file_id.split("_")[1])
                                        self.file_counter = max(self.file_counter, num)
                                except Exception as e:
                                    print(f"Error loading file {file_id}: {e}")
        except Exception as e:
            print(f"Error loading settings: {e}")
    
    def save_to_json(self):
        """Save uploaded files info ke JSON"""
        try:
            # Load existing settings
            settings = {}
            if os.path.exists("INVENTRA.json"):
                with open("INVENTRA.json", "r") as f:
                    settings = json.load(f)
            
            # Update uploaded_files section
            files_info = {}
            for file_id, file_data in self.uploaded_files.items():
                files_info[file_id] = {
                    'path': file_data['path'],
                    'size': file_data.get('size', 0),
                    'filename': file_data.get('filename', '')
                }
            
            settings["uploaded_files"] = files_info
            
            # Save to file
            with open("INVENTRA.json", "w") as f:
                json.dump(settings, f, indent=4)
                
        except Exception as e:
            print(f"Error saving to JSON: {e}")
        
    def show_splash_screen(self):
        """Show animated splash screen on startup"""
        self.splash = tk.Frame(self.root, bg=self.colors['primary'])
        self.splash.place(x=0, y=32, relwidth=1, relheight=1)
        self._lift_titlebar()
        
        logo_container = tk.Frame(self.splash, bg=self.colors['primary'])
        logo_container.pack(expand=True)
        
            # ===== LOAD LOGO =====
        logo = Image.open(resource_path("logo_trsp.png"))  # ganti dengan file logo kamu
        logo.thumbnail((500, 300), Image.LANCZOS)

        self.splash_logo = ImageTk.PhotoImage(logo)

        tk.Label(
            logo_container,
            image=self.splash_logo,
            bg=self.colors['primary']
        ).pack(pady=(0, 20))

        
        tk.Label(
            logo_container,
            text="An innovative inventory management solution",
            font=("Segoe UI", 20, "bold"),
            bg=self.colors['primary'],
            fg="white"
        ).pack()
        
        
        self.loading_label = tk.Label(
            logo_container,
            text="Loading application...",
            font=("Segoe UI", 11),
            bg=self.colors['primary'],
            fg="white"
        )
        self.loading_label.pack(pady=(0, 15))
        
        progress_container = tk.Frame(logo_container, bg=self.colors['primary'])
        progress_container.pack()
        
        progress_bg = tk.Canvas(
            progress_container,
            width=400,
            height=6,
            bg=self.colors['primary_dark'],
            highlightthickness=0
        )
        progress_bg.pack()
        
        self.progress_bar = progress_bg.create_rectangle(
            0, 0, 0, 6,
            fill="white",
            outline=""
        )

        # Flag sinkronisasi: tunggu keduanya selesai
        self._splash_anim_done = False
        self._splash_json_done = False

        # Jalankan pengecekan JSON di background thread
        threading.Thread(target=self._check_or_create_json, daemon=True).start()

        self.animate_progress(progress_bg, 0)

    # ── Default JSON ────────────────────────────────────────────────────────────

    DEFAULT_JSON = {
        "uploaded_files": {},
        "data_mappings": {
            "PLJM01":   {"source_file_id": None, "sheet_name": None, "columns": {
                "stock_code": "Stock Code", "rop": "ROP",
                "roq": "ROQ"
            }},
            "PLJM08":   {"source_file_id": None, "sheet_name": None, "columns": {
                "stock_code": "Stock Code", "item_name": "Item Name",
                "soh_akhir": "SOH Akhir"
            }},
            "SRD":      {"source_file_id": None, "sheet_name": None, "columns": {
                "stock_code": "STOCK_CODE", "qty_rcv_uop": "QTY_RCV_UOP",
                "creation_date": "CREATION_DATE"
            }},
            "SLN":      {"source_file_id": None, "sheet_name": None, "columns": {
                "stock_code": "Stock Code", "qty_req": "Qty Req",
                "qty_issued": "QTY_ISSUED", "last_acq_date": "Last Acq Date"
            }},
            "IR":       {"source_file_id": None, "sheet_name": None, "columns": {
                "stock_code": "Stock Code", "qty_req": "Qty Req",
                "qty_issued": "QTY_ISSUED", "req_by_date": "Req By Date"
            }},
            "PO":       {"source_file_id": None, "sheet_name": None, "columns": {
                "stock_code": "STOCK_CODE",   "supplier_name": "SUPPLIER_NAME",
                "order_date": "ORDER_DATE",   "receipt_status": "RECEIPT_STATUS",
                "qty_rcv_dir": "QTY_RCV_DIR", "curr_qty": "CURR_QTY"
            }},
            "LEVERING": {"source_file_id": None, "sheet_name": None, "columns": {
                "stock_code": "STOCK_CODE",  "receipt_status": "RECEIPT_STATUS",
                "curr_qty_p": "CURR_QTY_P", "due_site_date": "DUE_SITE_DATE"
            }},
            "ANALISIS SETTING": {"source_file_id": None, "sheet_name": None, "columns": {
                "supplier_name": "Supplier Name", "stock_code": "STOCK_CODE",
                "item_name": "Item Name",         "analisis": "Keterangan"
            }},
            "ANALISIS NON SETTING": {"source_file_id": None, "sheet_name": None, "columns": {
                "supplier_name": "Supplier Name", "stock_code": "STOCK_CODE",
                "item_name": "Item Name",         "analisis": "Keterangan"
            }}
        }
    }

    # ── JSON check / create ─────────────────────────────────────────────────────

    def _check_or_create_json(self):
        """
        Berjalan di background thread saat splash screen.
        - INVENTRA.json tidak ada  → buat baru dengan DEFAULT_JSON
        - INVENTRA.json sudah ada  → biarkan, tidak diubah
        Setelah selesai, cek apakah animasi splash juga sudah selesai,
        lalu lanjut ke menu.
        """
        try:
            if not os.path.exists("INVENTRA.json"):
                with open("INVENTRA.json", "w") as f:
                    json.dump(self.DEFAULT_JSON, f, indent=4)
                print("INVENTRA.json dibuat dengan nilai default.")
            else:
                print("INVENTRA.json sudah ada, tidak diubah.")
        except Exception as e:
            print(f"Error saat cek/buat INVENTRA.json: {e}")
        finally:
            self._splash_json_done = True
            self.root.after(0, self._try_finish_splash)

    def animate_progress(self, canvas, progress):
        """Animate the progress bar"""
        if progress <= 100:
            canvas.coords(self.progress_bar, 0, 0, progress * 4, 6)

            messages = [
                "Loading application...",
                "Initializing modules...",
                "Preparing interface...",
                "Setting up components...",
                "Almost ready..."
            ]

            message_index = min(int(progress / 20), len(messages) - 1)
            self.loading_label.config(text=messages[message_index])

            self.root.after(20, lambda: self.animate_progress(canvas, progress + 1))
        else:
            self._splash_anim_done = True
            self.root.after(300, self._try_finish_splash)

    def _try_finish_splash(self):
        """Lanjut ke menu hanya jika KEDUANYA sudah selesai (animasi + thread JSON)"""
        if self._splash_anim_done and self._splash_json_done:
            self.splash.destroy()
            self.show_module_chooser()

    def show_module_chooser(self):
        for w in self.root.winfo_children():
            try:
                if w is self._titlebar:
                    continue
                w.destroy()
            except Exception:
                pass

        c = self.colors
        self._chooser_frame = tk.Frame(self.root, bg=c['primary'])
        self._chooser_frame.place(x=0, y=32, relwidth=1, relheight=1)
        self._lift_titlebar()

        normal_color = c['primary']
        hover_color  = c['primary_dark']
        click_color  = self._darken_color(hover_color)

        # ── Exit pojok kanan atas ──
        # exit_btn = tk.Button(
        #     self._chooser_frame,
        #     text="✕ Exit",
        #     font=("Segoe UI", 11, "bold"),
        #     bg=normal_color, fg="white",
        #     activebackground=click_color,
        #     activeforeground="white",
        #     relief="flat", bd=0,
        #     cursor="hand2",
        #     command=self.root.quit,
        #     padx=20, pady=10
        # )
        # exit_btn.place(relx=1.0, x=-10, y=10, anchor="ne")
        # exit_btn.bind("<Enter>",           lambda e: exit_btn.config(bg=hover_color))
        # exit_btn.bind("<Leave>",           lambda e: exit_btn.config(bg=normal_color))
        # exit_btn.bind("<ButtonPress-1>",   lambda e: exit_btn.config(bg=click_color))
        # exit_btn.bind("<ButtonRelease-1>", lambda e: exit_btn.config(bg=hover_color))

        # ── Logo posisi tetap di tengah atas
        logo_frame = tk.Frame(self._chooser_frame, bg=c['primary'])
        logo_frame.place(relx=0.5, rely=0.30, anchor="center")

        try:
            logo = Image.open(resource_path("logo_trsp.png"))

            # Ambil ukuran layar
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()

            # Logo = 45% lebar layar, maksimal 50% tinggi layar
            max_w = int(screen_w * 0.35)
            max_h = int(screen_h * 0.23)

            # Hitung rasio agar proporsional
            orig_w, orig_h = logo.size
            ratio = min(max_w / orig_w, max_h / orig_h)
            new_w = int(orig_w * ratio)
            new_h = int(orig_h * ratio)

            logo = logo.resize((new_w, new_h), Image.LANCZOS)
            self._chooser_logo = ImageTk.PhotoImage(logo)
            tk.Label(logo_frame, image=self._chooser_logo,
                    bg=c['primary']).pack()
        except Exception:
            # Fallback teks dengan ukuran font dinamis
            screen_w = self.root.winfo_screenwidth()
            font_size = max(32, int(screen_w * 0.03))
            tk.Label(logo_frame, text="INVENTRA",
                    font=("Segoe UI", font_size, "bold"),
                    bg=c['primary'], fg="white").pack()

        # ── Tombol menu dipisah, posisi lebih ke bawah
        btn_frame = tk.Frame(self._chooser_frame, bg=c['primary'])
        btn_frame.place(relx=0.5, rely=0.63, anchor="center")

        def make_canvas_btn(parent, icon_file, label, command):
            W, H, R = 220, 220, 22
            ICON_SIZE = 120
            LABEL_H   = 45

            canvas = tk.Canvas(parent, width=W, height=H,
                               bg=c['primary'], highlightthickness=0,
                               cursor="hand2")
            canvas.pack(side="left", padx=35)

            try:
                raw = Image.open(resource_path(icon_file)).convert("RGBA")
                raw.thumbnail((ICON_SIZE, ICON_SIZE), Image.LANCZOS)
                photo = ImageTk.PhotoImage(raw)
            except Exception:
                photo = None

            canvas._photo = photo

            def draw_btn(fill):
                canvas.delete("all")
                # Rounded rect
                canvas.create_arc(0,       0,       2*R,   2*R,   start=90,  extent=90,  fill=fill, outline=fill)
                canvas.create_arc(W-2*R,   0,       W,     2*R,   start=0,   extent=90,  fill=fill, outline=fill)
                canvas.create_arc(0,       H-2*R,   2*R,   H,     start=180, extent=90,  fill=fill, outline=fill)
                canvas.create_arc(W-2*R,   H-2*R,   W,     H,     start=270, extent=90,  fill=fill, outline=fill)
                canvas.create_rectangle(R,   0,  W-R,  H,   fill=fill, outline=fill)
                canvas.create_rectangle(0,   R,  W,    H-R, fill=fill, outline=fill)
                # Icon
                icon_area_h = H - LABEL_H
                if canvas._photo:
                    canvas.create_image(W//2, icon_area_h//2, image=canvas._photo, anchor="center")
                else:
                    canvas.create_text(W//2, icon_area_h//2,
                                       text="⚙" if "setting" in icon_file else "📋",
                                       font=("Segoe UI", 48), fill=c['primary'], anchor="center")
                # Label
                canvas.create_text(W//2, H - LABEL_H//2,
                                   text=label,
                                   font=("Segoe UI", 18, "bold"),
                                   fill=c['primary'], anchor="center")

            draw_btn("white")
            canvas.bind("<Enter>",           lambda e: draw_btn("#f3f4f6"))
            canvas.bind("<Leave>",           lambda e: draw_btn("white"))
            canvas.bind("<ButtonPress-1>",   lambda e: draw_btn("#e5e7eb"))
            canvas.bind("<ButtonRelease-1>", lambda e: (draw_btn("#f3f4f6"), command()))

        make_canvas_btn(btn_frame, "icon_settingan.png", "Settingan",  self._open_inventra)
        make_canvas_btn(btn_frame, "icon_drp.png",       "Proses DRP", self._open_drp)

    def _back_to_chooser(self):
        try:
            self.main_canvas.unbind_all("<MouseWheel>")
        except Exception:
            pass
        
        for w in self.root.winfo_children():
            try:
                if w is self._titlebar:
                    continue
                w.destroy()
            except Exception:
                pass
        self.show_module_chooser()

    def _open_inventra(self):
        if hasattr(self, "_chooser_frame"):
            self._chooser_frame.destroy()
        self._show_loading_screen(
            message="Memuat data Settingan...",
            thread_target=self._load_inventra_thread
        )

    def _load_inventra_thread(self):
        """Load INVENTRA.json + file Excel di background"""
        try:
            self.load_from_json()
        except Exception as e:
            print("Error loading inventra:", e)
        self.root.after(0, self._finish_inventra)

    def _finish_inventra(self):
        self._hide_loading_screen()
        self.setup_ui()

    def _open_drp(self):
        if hasattr(self, "_chooser_frame"):
            self._chooser_frame.destroy()
        self._show_loading_screen(
            message="Menghubungkan ke Google Sheets...",
            thread_target=self._load_drp_thread
        )

    def _load_drp_thread(self):
        """Cek koneksi Google Sheets di background"""
        import urllib.request
        try:
            req = urllib.request.Request(
                WEB_APP_URL,
                headers={"Content-Type": "application/json"}
            )
            urllib.request.urlopen(req, timeout=10)
            self._drp_conn_ok = True
        except Exception as e:
            print("GSheet connection:", e)
            self._drp_conn_ok = False
        self.root.after(0, self._finish_drp)

    def _finish_drp(self):
        self._hide_loading_screen()

        def back():
            for w in self.root.winfo_children():
                try:
                    if w is self._titlebar:
                        continue
                    w.destroy()
                except Exception:
                    pass
            self.show_module_chooser()

        DRPApp(root=self.root, colors=self.colors, on_back=back,
               lift_titlebar=self._lift_titlebar)

    # ── Loading Screen ─────────────────────────────────────────────────────────

    def _show_loading_screen(self, message, thread_target):
        """Tampilkan loading screen, lalu jalankan thread_target di background"""
        self._loading_frame = tk.Frame(self.root, bg=self.colors['primary'])
        self._loading_frame.place(x=0, y=32, relwidth=1, relheight=1)
        self._lift_titlebar()

        container = tk.Frame(self._loading_frame, bg=self.colors['primary'])
        container.place(relx=0.5, rely=0.5, anchor="center")

        # Logo
        try:
            logo = Image.open(resource_path("logo_trsp.png"))
            logo.thumbnail((500, 300), Image.LANCZOS)
            self._loading_logo = ImageTk.PhotoImage(logo)
            tk.Label(
                container,
                image=self._loading_logo,
                bg=self.colors['primary']
            ).pack(pady=(0, 30))
        except Exception:
            tk.Label(
                container,
                text="INVENTRA",
                font=("Segoe UI", 32, "bold"),
                bg=self.colors['primary'],
                fg="white"
            ).pack(pady=(0, 30))

        # Pesan
        self._loading_msg_label = tk.Label(
            container,
            text=message,
            font=("Segoe UI", 13),
            bg=self.colors['primary'],
            fg="white"
        )
        self._loading_msg_label.pack(pady=(0, 20))

        # Progress bar animasi
        prog_bg = tk.Canvas(
            container,
            width=360, height=6,
            bg=self.colors['primary_dark'],
            highlightthickness=0
        )
        prog_bg.pack()
        self._loading_bar = prog_bg.create_rectangle(0, 0, 0, 6, fill="white", outline="")
        self._loading_canvas = prog_bg
        self._loading_progress = 0
        self._animate_loading_bar()

        # Jalankan thread
        thread = threading.Thread(target=thread_target, daemon=True)
        thread.start()

    def _animate_loading_bar(self):
        """Animasi indeterminate loading bar (bolak-balik)"""
        if not hasattr(self, "_loading_canvas") or not self._loading_canvas.winfo_exists():
            return

        self._loading_progress = (self._loading_progress + 3) % 120
        p = self._loading_progress
        # Efek sliding bar: panjang 120px, gerak dari 0 ke 360
        start_x = max(0, p * 3)
        end_x   = min(360, start_x + 120)
        self._loading_canvas.coords(self._loading_bar, start_x, 0, end_x, 6)
        self._loading_anim_id = self.root.after(16, self._animate_loading_bar)

    def _hide_loading_screen(self):
        """Hentikan animasi dan hapus loading screen"""
        if hasattr(self, "_loading_anim_id"):
            self.root.after_cancel(self._loading_anim_id)
        if hasattr(self, "_loading_frame") and self._loading_frame.winfo_exists():
            self._loading_frame.destroy()
        
    def setup_ui(self):
        # Main container with scrollbar
        self.main_canvas = tk.Canvas(self.root, bg=self.colors['bg'], highlightthickness=0)
        main_scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.main_canvas.yview)
        
        self.scrollable_frame = tk.Frame(self.main_canvas, bg=self.colors['bg'])
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
        )
        
        self.canvas_window = self.main_canvas.create_window(
            (0, 0),
            window=self.scrollable_frame,
            anchor="nw"
        )

        self.main_canvas.bind(
            "<Configure>",
            lambda e: self.main_canvas.itemconfig(
                self.canvas_window,
                width=e.width
            )
        )
        self.main_canvas.configure(yscrollcommand=main_scrollbar.set)
        
        self.main_canvas.pack(side="left", fill="both", expand=True)
        main_scrollbar.pack(side="right", fill="y")
        self._lift_titlebar()
        
        # Bind mousewheel
        self.main_canvas.bind("<Enter>", self._bind_mousewheel)
        self.main_canvas.bind("<Leave>", self._unbind_mousewheel)
        
        # Header
        self.create_header()
        
        # Upload Section
        self.create_upload_section()
        
        # Action buttons
        self.create_action_buttons()
        
        # Data Preview Section
        self.create_data_preview()
        
    def create_header(self):
        header_frame = tk.Frame(self.scrollable_frame, bg=self.colors['primary'], height=130)
        header_frame.pack(fill="x", side="top")
        header_frame.pack_propagate(False)
        
        header_content = tk.Frame(header_frame, bg=self.colors['primary'])
        header_content.pack(expand=True, fill="both", padx=40, pady=20)
        
        left_header = tk.Frame(header_content, bg=self.colors['primary'])
        left_header.pack(side="left", fill="y")

        # ===== LOAD LOGO =====
        logo = Image.open(resource_path("logo_trsp.png"))  # ganti dengan nama file kamu
        logo.thumbnail((300, 120), Image.LANCZOS)

        self.header_logo = ImageTk.PhotoImage(logo)

        tk.Label(
            left_header,
            image=self.header_logo,
            bg=self.colors['primary']
        ).pack(anchor="w")

        
        right_header = tk.Frame(header_content, bg=self.colors['primary'])
        right_header.pack(side="right", anchor="e")
        
        # === EXIT BUTTON MODERN STATE ===
        normal_color = self.colors['primary']               # warna header
        hover_color = self.colors['primary_dark']          # saat hover
        click_color = self._darken_color(hover_color)      # lebih gelap lagi

        tutorial_btn = tk.Button(
            right_header,
            text="❓ Help",
            font=("Segoe UI", 11, "bold"),
            bg=normal_color,
            fg="white",
            activebackground=click_color,   # supaya tidak jadi putih saat klik
            activeforeground="white",
            relief="flat",
            cursor="hand2",
            command=self.open_tutorial,
            bd=0,
            padx=25,
            pady=10
        )
        tutorial_btn.pack(side="left", padx=(0, 10))

        # Hover
        tutorial_btn.bind("<Enter>", lambda e: tutorial_btn.config(bg=hover_color))

        # Leave
        tutorial_btn.bind("<Leave>", lambda e: tutorial_btn.config(bg=normal_color))

        # Saat ditekan
        tutorial_btn.bind("<ButtonPress-1>", lambda e: tutorial_btn.config(bg=click_color))

        # Saat dilepas
        tutorial_btn.bind("<ButtonRelease-1>", lambda e: tutorial_btn.config(bg=hover_color))


        back_btn = tk.Button(
            right_header,
            text="← Kembali",
            font=("Segoe UI", 11, "bold"),
            bg=normal_color, fg="white",
            activebackground=click_color,
            activeforeground="white",
            relief="flat", bd=0,
            highlightthickness=0,
            cursor="hand2",
            command=self._back_to_chooser,
            padx=25, pady=10
        )
        back_btn.pack(side="left")
        back_btn.bind("<Enter>",           lambda e: back_btn.config(bg=hover_color))
        back_btn.bind("<Leave>",           lambda e: back_btn.config(bg=normal_color))
        back_btn.bind("<ButtonPress-1>",   lambda e: back_btn.config(bg=click_color))
        back_btn.bind("<ButtonRelease-1>", lambda e: back_btn.config(bg=hover_color))

    def open_tutorial(self):
        TutorialScreen(self.root, self.colors, None)
        
    def create_upload_section(self):
        """Create file upload section"""
        upload_frame = tk.Frame(self.scrollable_frame, bg=self.colors['bg'])
        upload_frame.pack(fill="both", expand=True, padx=40, pady=(20, 10))
        
        # Title
        title_frame = tk.Frame(upload_frame, bg=self.colors['bg'])
        title_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(
            title_frame,
            text="Upload Files",
            font=("Segoe UI", 24, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['text'],
            anchor="w"
        ).pack(side="left")
        
        self.file_count_label = tk.Label(
            title_frame,
            text="0 files",
            font=("Segoe UI", 11),
            bg=self.colors['bg'],
            fg=self.colors['text_light']
        )
        self.file_count_label.pack(side="left", padx=(15, 0))
        
        # Upload card container
        self.file_cards_frame = tk.Frame(upload_frame, bg=self.colors['bg'])
        self.file_cards_frame.pack(fill="both", expand=True)
        
        # Empty state or file cards
        self.file_cards = {}
        
        if not self.uploaded_files:
            self.show_empty_state()
        else:
            # Recreate cards from loaded files
            for file_id, file_data in self.uploaded_files.items():
                self.create_file_card(file_id, file_data)
            self.update_file_count()
        
    def show_empty_state(self):
        """Show empty state when no files uploaded"""
        if hasattr(self, 'empty_state_frame'):
            self.empty_state_frame.destroy()
            
        self.empty_state_frame = tk.Frame(self.file_cards_frame, bg=self.colors['card_bg'], relief="solid", bd=1)
        self.empty_state_frame.pack(fill="both", expand=True, pady=10)
        
        content = tk.Frame(self.empty_state_frame, bg=self.colors['card_bg'])
        content.pack(expand=True, pady=80)
        
        tk.Label(
            content,
            text="📁",
            font=("Segoe UI", 48),
            bg=self.colors['card_bg']
        ).pack()
        
        tk.Label(
            content,
            text="No files uploaded yet",
            font=("Segoe UI", 16, "bold"),
            bg=self.colors['card_bg'],
            fg=self.colors['text']
        ).pack(pady=(20, 10))
        
        tk.Label(
            content,
            text="Please upload your Excel files to get started",
            font=("Segoe UI", 11),
            bg=self.colors['card_bg'],
            fg=self.colors['text_light']
        ).pack()
    
    def create_file_card(self, file_id, file_data):
        """Create a card for uploaded file"""
        if hasattr(self, 'empty_state_frame'):
            self.empty_state_frame.destroy()
        
        card = tk.Frame(self.file_cards_frame, bg=self.colors['card_bg'], relief="solid", bd=1)
        card.pack(fill="x", pady=5)
        self.add_card_shadow(card)
        
        card_content = tk.Frame(card, bg=self.colors['card_bg'])
        card_content.pack(fill="x", padx=20, pady=15)
        
        # Left: File info
        left_frame = tk.Frame(card_content, bg=self.colors['card_bg'])
        left_frame.pack(side="left", fill="x", expand=True)
        
        tk.Label(
            left_frame,
            text=f"📄 {file_data['filename']}",
            font=("Segoe UI", 12, "bold"),
            bg=self.colors['card_bg'],
            fg=self.colors['text'],
            anchor="w"
        ).pack(anchor="w")
        
        info_text = f"{len(file_data['sheets'])} sheets • {file_data['size'] / 1024:.1f} KB"
        tk.Label(
            left_frame,
            text=info_text,
            font=("Segoe UI", 9),
            bg=self.colors['card_bg'],
            fg=self.colors['text_light'],
            anchor="w"
        ).pack(anchor="w", pady=(3, 0))
        
        # Right: Delete button
        delete_btn = tk.Button(
            card_content,
            text="✕",
            font=("Segoe UI", 12, "bold"),
            bg=self.colors['card_bg'],
            fg=self.colors['text_light'],
            relief="flat",
            cursor="hand2",
            command=lambda: self.delete_file(file_id),
            width=3
        )
        delete_btn.pack(side="right")
        
        delete_btn.bind("<Enter>", lambda e: delete_btn.config(fg=self.colors['primary'], bg=self.colors['primary_light']))
        delete_btn.bind("<Leave>", lambda e: delete_btn.config(fg=self.colors['text_light'], bg=self.colors['card_bg']))
        
        self.file_cards[file_id] = {
            'frame': card,
            'data': file_data
        }
    
    def upload_file(self):
        file_paths = filedialog.askopenfilenames(
            title="Select Excel Files",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if not file_paths:
            return

        # 🔒 Disable UI sementara
        self.disable_ui_during_process()

        # 🚀 Jalankan di thread
        thread = threading.Thread(
            target=self._upload_file_thread,
            args=(file_paths,),
            daemon=True
        )
        thread.start()

    def _upload_file_thread(self, file_paths):
        """Proses berat baca Excel di background"""
        new_files = {}

        for file_path in file_paths:
            try:
                excel_file = pd.ExcelFile(file_path)
                sheets_data = {}

                for sheet_name in excel_file.sheet_names:
                    df = self.read_excel_auto_header(file_path, sheet_name)
                    sheets_data[sheet_name] = df

                self.file_counter += 1
                file_id = f"file_{self.file_counter}"

                import os as os_module
                file_size = os_module.path.getsize(file_path)
                filename = os_module.path.basename(file_path)

                new_files[file_id] = {
                    'path': file_path,
                    'sheets': sheets_data,
                    'size': file_size,
                    'filename': filename
                }

            except Exception as e:
                print("Upload error:", e)

        # 🔥 Kembali ke main thread untuk update UI
        self.root.after(0, lambda: self._finish_upload(new_files))

    def _finish_upload(self, new_files):
        """Update UI setelah upload selesai"""

        for file_id, file_data in new_files.items():
            self.uploaded_files[file_id] = file_data
            self.create_file_card(file_id, file_data)

        self.update_file_count()
        self.save_to_json()

        # 🔓 Aktifkan lagi UI
        self.enable_ui_after_process()

    def disable_ui_during_process(self):
        """Nonaktifkan tombol & klik saat proses berjalan"""
        self.root.config(cursor="watch")

        for child in self.root.winfo_children():
            try:
                child.configure(state="disabled")
            except:
                pass

    def enable_ui_after_process(self):
        """Aktifkan kembali UI"""
        self.root.config(cursor="")

        for child in self.root.winfo_children():
            try:
                child.configure(state="normal")
            except:
                pass
    
    def delete_file(self, file_id):
        """Delete uploaded file"""
        if messagebox.askyesno("Confirm", "Delete this file?"):
            # Remove from uploaded_files
            if file_id in self.uploaded_files:
                del self.uploaded_files[file_id]
            
            # Remove card
            if file_id in self.file_cards:
                self.file_cards[file_id]['frame'].destroy()
                del self.file_cards[file_id]
            
            # Update count
            self.update_file_count()
            
            # Save to JSON
            self.save_to_json()
            
            # Show empty state if no files
            if not self.uploaded_files:
                self.show_empty_state()
    
    def update_file_count(self):
        """Update file count label"""
        count = len(self.uploaded_files)
        self.file_count_label.config(text=f"{count} file{'s' if count != 1 else ''}")
    
    def create_action_buttons(self):
        """Create action buttons section"""
        actions_frame = tk.Frame(self.scrollable_frame, bg=self.colors['bg'])
        actions_frame.pack(fill="x", padx=40, pady=20)
        
        button_container = tk.Frame(actions_frame, bg=self.colors['bg'])
        button_container.pack()
        
        # Upload button
        upload_btn = self.create_modern_button(
            button_container,
            "+ Upload File",
            self.upload_file,
            self.colors['primary']
        )
        upload_btn.pack(side="left", padx=5)
        
        # Process button
        process_btn = self.create_modern_button(
            button_container,
            "⚙ Process",
            self.open_mapping_dialog,
            self.colors['success']
        )
        process_btn.pack(side="left", padx=5)
        
        # Save button
        save_btn = self.create_modern_button(
            button_container,
            "💾 Save Excel",
            self.save_excel,
            self.colors['warning']
        )
        save_btn.pack(side="left", padx=5)
        
        # Reset button
        reset_btn = self.create_modern_button(
            button_container,
            "🔄 Reset All",
            self.reset_all,
            self.colors['text_light']
        )
        reset_btn.pack(side="left", padx=5)
    
    def create_data_preview(self):
        """Create data preview section"""
        preview_frame = tk.Frame(self.scrollable_frame, bg=self.colors['bg'])
        preview_frame.pack(fill="both", expand=True, padx=40, pady=(10, 40))
        
        # Title
        title_frame = tk.Frame(preview_frame, bg=self.colors['bg'])
        title_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(
            title_frame,
            text="Data Preview",
            font=("Segoe UI", 24, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['text']
        ).pack(side="left")
        
        # View selector buttons
        view_selector_frame = tk.Frame(preview_frame, bg=self.colors['bg'])
        view_selector_frame.pack(fill="x", pady=(0, 10))

        self.view_buttons_parent = view_selector_frame
        self.view_buttons = {}

        def place_buttons():
            # Clear dulu semua widget
            for widget in view_selector_frame.winfo_children():
                widget.place_forget()

            view_selector_frame.update_idletasks()
            max_width = view_selector_frame.winfo_width()
            if max_width <= 1:
                max_width = preview_frame.winfo_width() - 80  # fallback

            x, y = 0, 0
            row_height = 0
            pad_x, pad_y = 5, 5

            for key, btn in self.view_buttons.items():
                btn.update_idletasks()
                bw = btn.winfo_reqwidth()
                bh = btn.winfo_reqheight()

                # Kalau melebihi lebar, pindah ke baris bawah
                if x + bw > max_width and x > 0:
                    x = 0
                    y += row_height + pad_y
                    row_height = 0

                btn.place(x=x, y=y)
                x += bw + pad_x
                row_height = max(row_height, bh)

            # Set tinggi frame sesuai konten
            total_height = y + row_height + pad_y
            view_selector_frame.config(height=max(total_height, 35))

        self._place_view_buttons = place_buttons

        for key in self.data_store.keys():
            if self.data_store.get(key) is None:
                continue

            btn = self.create_view_button(view_selector_frame, key, key)
            self.view_buttons[key] = btn

        # Jalankan setelah render
        view_selector_frame.after(100, place_buttons)

        # Re-wrap kalau window di-resize
        preview_frame.bind("<Configure>", lambda e: place_buttons())

        
        # for key, label in views:
        #     btn = self.create_view_button(view_selector_frame, key, label)
        #     btn.pack(side="left", padx=(0, 5))
        #     self.view_buttons[key] = btn
        
        # Treeview container
        tree_container = tk.Frame(preview_frame, bg=self.colors['card_bg'], relief="solid", bd=1)
        tree_container.pack(fill="both", expand=True)
        
        # Treeview with scrollbars
        tree_scroll_y = ttk.Scrollbar(tree_container, orient="vertical")
        tree_scroll_x = ttk.Scrollbar(tree_container, orient="horizontal")
        
        self.tree = ttk.Treeview(
            tree_container,
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set,
            show="headings",
            height=15
        )
        
        tree_scroll_y.config(command=self.tree.yview)
        tree_scroll_x.config(command=self.tree.xview)
        
        tree_scroll_y.pack(side="right", fill="y")
        tree_scroll_x.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)
        
        # Info label
        self.info_label = tk.Label(
            preview_frame,
            text="No data to display",
            font=("Segoe UI", 10),
            bg=self.colors['bg'],
            fg=self.colors['text_light']
        )
        self.info_label.pack(pady=(10, 0))
    
    def open_mapping_dialog(self):
        """Open mapping configuration dialog"""
        if not self.uploaded_files:
            messagebox.showwarning("No Files", "Please upload files first.")
            return

        # ===== CREATE DARK OVERLAY =====
        self.overlay = tk.Toplevel(self.root)
        self.overlay.overrideredirect(True)
        self.overlay.attributes("-alpha", 0.8)
        self.overlay.configure(bg="black")

        # samakan ukuran dengan window utama
        self.overlay.geometry(
            f"{self.root.winfo_width()}x{self.root.winfo_height()}+"
            f"{self.root.winfo_rootx()}+{self.root.winfo_rooty()}"
        )

        self.overlay.lift()
        self.overlay.transient(self.root)
        def keep_overlay_on_top(event=None):
            if self.overlay and self.overlay.winfo_exists():
                self.overlay.lift()
                self.overlay.attributes("-topmost", True)
                self.overlay.attributes("-topmost", False)

        self.root.bind("<FocusIn>", keep_overlay_on_top)


        # buka dialog mapping
        MappingDialog(
            self.root,
            self.colors,
            self.uploaded_files,
            self.process_data_with_mapping,
            overlay=self.overlay   # <-- kirim overlay
        )

    
    def process_data_with_mapping(self, mappings):
        """Process data setelah mapping dikonfigurasi"""
        try:
            # Map data dari uploaded files
            for data_type, mapping in mappings.items():
                file_id = mapping['source_file_id']
                sheet_name = mapping['sheet_name']
                
                if file_id in self.uploaded_files:
                    if sheet_name in self.uploaded_files[file_id]['sheets']:
                        self.data_store[data_type] = self.uploaded_files[file_id]['sheets'][sheet_name].copy()
                    else:
                        messagebox.showwarning(
                            "Sheet Not Found",
                            f"{data_type}: Sheet '{sheet_name}' not found"
                        )
                        return
                else:
                    messagebox.showwarning(
                        "File Not Found",
                        f"{data_type}: File '{file_id}' not uploaded"
                    )
                    return
            
            # Check if all data is available
            REQUIRED_SECTIONS = ["PLJM08", "SRD", "SLN", "IR", "PO", "LEVERING", "ANALISIS SETTING", "ANALISIS NON SETTING"]
            missing = [key for key in REQUIRED_SECTIONS if self.data_store.get(key) is None]
            if missing:
                messagebox.showwarning(
                    "Missing Data",
                    f"Missing data for: {', '.join(missing)}"
                )
                return
            
            # Process data
            self.data_store = data.filter_all(
                self.data_store,
                mappings   # ← ini mapping dari dialog
            )
            
            messagebox.showinfo("Success", "Data processed successfully!")
            
            # Show PLJM08 preview
            if self.data_store.get("PLJM08") is not None:
                self.show_data("PLJM08")
            
            # Scroll to preview section
            self.root.after(100, self.scroll_to_preview)

            self.refresh_view_buttons()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to process data:\n{str(e)}")

    def refresh_view_buttons(self):
        for btn in self.view_buttons.values():
            btn.destroy()

        self.view_buttons.clear()

        parent = self.view_buttons_parent  # simpan parent saat create_data_preview

        for key, df in self.data_store.items():
            if df is None:
                continue
            if key == "PLJM01":
                continue
            
            btn = self.create_view_button(parent, key, key)
            btn.pack(side="left", padx=(0, 5))
            self.view_buttons[key] = btn
    
    def scroll_to_preview(self):
        """Scroll to data preview section"""
        self.main_canvas.yview_moveto(1.0)
    
    def add_card_shadow(self, widget):
        """Add shadow effect to card"""
        widget.configure(relief="flat", bd=0)
    
    def _bind_mousewheel(self, event):
        self.main_canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _unbind_mousewheel(self, event):
        self.main_canvas.unbind_all("<MouseWheel>")

    
    def _on_mousewheel(self, event):
        """Handle mousewheel scrolling"""
        if not hasattr(self, 'main_canvas') or not self.main_canvas.winfo_exists():
            return
        self.main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    
    def create_modern_button(self, parent, text, command, bg_color):
        """Create modern styled button"""
        btn = tk.Button(
            parent,
            text=text,
            font=("Segoe UI", 11, "bold"),
            bg=bg_color,
            fg="white",
            relief="flat",
            cursor="hand2",
            command=command,
            padx=25,
            pady=12
        )
        
        darker_color = self._darken_color(bg_color)
        btn.bind("<Enter>", lambda e: btn.config(bg=darker_color))
        btn.bind("<Leave>", lambda e: btn.config(bg=bg_color))
        
        return btn
    
    def _darken_color(self, color):
        """Darken a hex color"""
        if color.startswith('#'):
            color = color[1:]
        r, g, b = int(color[:2], 16), int(color[2:4], 16), int(color[4:], 16)
        r, g, b = max(0, r-30), max(0, g-30), max(0, b-30)
        return f"#{r:02x}{g:02x}{b:02x}"
    
    def create_view_button(self, parent, key, label):
        """Create view selector button"""
        btn = tk.Button(
            parent,
            text=label,
            font=("Segoe UI", 9),
            bg=self.colors['border'],
            fg=self.colors['text'],
            relief="flat",
            cursor="hand2",
            command=lambda: self.show_data(key),
            padx=12,
            pady=6
        )
        
        btn.bind("<Enter>", lambda e: btn.config(bg=self.colors['primary_light']) if self.current_view != key else None)
        btn.bind("<Leave>", lambda e: btn.config(bg=self.colors['border']) if self.current_view != key else None)
        
        return btn
    
    def show_data(self, data_key):
        """Display data in treeview"""
        if data_key not in self.data_store or self.data_store[data_key] is None:
            self.info_label.config(text=f"No data available for {data_key}")
            return

        self.current_view = data_key

        for key, btn in self.view_buttons.items():
            if key == data_key:
                btn.config(bg=self.colors['primary'], fg="white")
            else:
                btn.config(bg=self.colors['border'], fg=self.colors['text'])

        self.tree.delete(*self.tree.get_children())

        df = self.data_store[data_key]

        self.tree["columns"] = list(df.columns)
        self.tree["show"] = "headings"

        # 🔥 AUTO WIDTH CALCULATION
        for col in df.columns:
            self.tree.heading(col, text=col)

            try:
                max_length = len(str(col))

                for value in df[col]:
                    value_length = len(str(value))
                    if value_length > max_length:
                        max_length = value_length

            except Exception:
                max_length = 10  # fallback aman

            col_width = int(max_length * 7 + 30)

            self.tree.column(
                col,
                width=col_width,
                minwidth=100,
                anchor="w",
                stretch=False
            )



        for _, row in df.iterrows():
            values = [
                "" if pd.isna(val) else str(val)
                for val in row
            ]
            self.tree.insert("", "end", values=values)

        self.info_label.config(
            text=f"Showing: {data_key} ({len(df)} rows, {len(df.columns)} columns)"
        )

    
    # ==============================
    # PUBLIC METHOD
    # ==============================
    def save_excel(self):
        """Save processed data to Excel (thread-safe & UI-safe)"""

        if not self.data_store or self.data_store.get("PLJM08") is None:
            messagebox.showwarning("No Data", "Please process data first.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"INVENTRA_{datetime.now().strftime('%d %B %Y_%H-%M')}.xlsx"
        )

        if not file_path:
            return

        # Copy data to avoid mutation during thread execution
        # PLJM01 tidak disertakan di Excel output
        data_copy = {
            k: v.copy() if isinstance(v, pd.DataFrame) else v
            for k, v in self.data_store.items()
            if k not in ("PLJM01",)
        }

        # Ambil tanggal previous result dari sheet COVER cell C7
        analisis_file_date = None
        try:
            ns_file_id = None
            if os.path.exists("INVENTRA.json"):
                with open("INVENTRA.json", "r") as f:
                    _cfg = json.load(f)
                ns_file_id = _cfg.get("data_mappings", {}).get("ANALISIS NON SETTING", {}).get("source_file_id")

            if ns_file_id and ns_file_id in self.uploaded_files:
                ns_file_path = self.uploaded_files[ns_file_id].get("path")
                if ns_file_path and os.path.exists(ns_file_path):
                    from openpyxl import load_workbook
                    wb_prev = load_workbook(ns_file_path, read_only=True, data_only=True)
                    cover_sheet_name = next(
                        (s for s in wb_prev.sheetnames if "COVER" in s.upper()), None
                    )
                    if cover_sheet_name:
                        ws_prev = wb_prev[cover_sheet_name]
                        cell_val = ws_prev["C7"].value
                        if isinstance(cell_val, datetime):
                            analisis_file_date = cell_val
                        elif cell_val:
                            # Format: '30 March 2026  |  14:53 WIB' — ambil bagian sebelum '|'
                            raw = str(cell_val).split("|")[0].strip()
                            analisis_file_date = pd.to_datetime(raw, dayfirst=True, errors="coerce")
                            if pd.isna(analisis_file_date):
                                analisis_file_date = None
                    wb_prev.close()
        except Exception as e:
            print(f"[WARN] Gagal baca tanggal previous result: {e}")
            analisis_file_date = None

        # Disable save button if exists
        if hasattr(self, "btn_save"):
            self.btn_save.config(state="disabled")

        self.show_processing_overlay("Saving Excel...\nPlease wait")

        thread = threading.Thread(
            target=self._save_excel_thread,
            args=(file_path, data_copy, analisis_file_date),
            daemon=True
        )
        thread.start()

    # ==============================
    # THREAD WORKER
    # ==============================
    def _save_excel_thread(self, file_path, data_store, analisis_file_date=None):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # ── palette (sesuai tema INVENTRA) ──────────────────
            RED_DARK  = "7F1D1D"
            RED_MAIN  = "DC2626"
            RED_LIGHT = "FEE2E2"
            RED_MID   = "FECACA"
            WHITE     = "FFFFFF"
            GRAY_BG   = "F3F4F6"
            GRAY_LT   = "F9FAFB"
            GRAY_TEXT = "374151"
            GRAY_MID  = "6B7280"

            # ── helpers ─────────────────────────────────────────
            def _side(c="D1D5DB", s="thin"):
                return Side(style=s, color=c)

            def _border(c="D1D5DB"):
                s = _side(c)
                return Border(left=s, right=s, top=s, bottom=s)

            def _fill(c):
                return PatternFill("solid", fgColor=c)

            def _font(bold=False, size=10, color="000000", italic=False):
                return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

            def _align(h="left", v="center", wrap=False):
                return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

            def fill_row(ws, row, color, ncols=8):
                for col in range(1, ncols + 1):
                    ws.cell(row=row, column=col).fill = _fill(color)

            def mset(ws, r, c, re, ce, value="", bold=False, size=10,
                     color="000000", bg=None, halign="left", italic=False):
                ws.merge_cells(start_row=r, start_column=c, end_row=re, end_column=ce)
                cell = ws.cell(row=r, column=c, value=value)
                cell.font      = _font(bold, size, color, italic)
                cell.alignment = _align(halign)
                if bg:
                    cell.fill = _fill(bg)
                return cell

            # ── COVER SHEET ──────────────────────────────────────
            def write_cover(ws, ds, now=None):
                from openpyxl.worksheet.page import PageMargins
                if now is None:
                    now = datetime.now()
                ws.sheet_view.showGridLines     = False
                ws.sheet_view.showRowColHeaders = False

                # ── Landscape A4 ──
                ws.page_setup.orientation    = "landscape"
                ws.page_setup.paperSize       = ws.PAPERSIZE_A4
                ws.page_setup.fitToPage       = True
                ws.page_setup.fitToWidth      = 1
                ws.page_setup.fitToHeight     = 0
                ws.page_margins               = PageMargins(
                    left=0.5, right=0.5, top=0.5, bottom=0.5
                )

                # ── kolom: A=margin, B=No, C=Sheet Name(lebar), D=Rows, E=Cols, F=margin ──
                # Total 6 kolom, lebih lebar & proporsional
                col_w = {1: 3, 2: 22, 3: 35, 4: 16, 5: 16, 6: 3}
                for col, w in col_w.items():
                    ws.column_dimensions[get_column_letter(col)].width = w

                NCOLS = 6   # kolom aktif A-F

                valid = [(k, v) for k, v in ds.items() if isinstance(v, pd.DataFrame)]

                # row heights
                rh = {
                    1: 15,   # top margin
                    2: 65,   # INVENTRA title
                    3: 22,   # subtitle
                    4: 12,   # accent stripe
                    5: 14,   # spacer
                    6: 24,   # period
                    7: 20,   # generated date
                    8: 14,   # spacer
                    9: 26,   # CONTENTS heading
                    10: 12,  # spacer
                    11: 28,  # table header
                }
                for i in range(len(valid)):
                    rh[12 + i] = 22
                rh[12 + len(valid)] = 12   # spacer
                rh[13 + len(valid)] = 20   # footer
                for r, h in rh.items():
                    ws.row_dimensions[r].height = h

                total_rows = 14 + len(valid)

                # BG seluruh area
                for row in range(1, total_rows + 1):
                    fill_row(ws, row, GRAY_BG, ncols=NCOLS)

                # header gelap rows 1-4
                for row in range(1, 5):
                    fill_row(ws, row, RED_DARK, ncols=NCOLS)

                mset(ws, 2, 2, 2, 5, "INVENTRA",
                     bold=True, size=40, color=WHITE, bg=RED_DARK, halign="center")
                mset(ws, 3, 2, 3, 5, "Inventory Management System  —  Export Report",
                     bold=False, size=12, color=RED_MID, bg=RED_DARK,
                     halign="center", italic=True)

                # accent stripe
                fill_row(ws, 4, RED_MAIN, ncols=NCOLS)

                # spacer
                fill_row(ws, 5, WHITE, ncols=NCOLS)

                # period & generated
                fill_row(ws, 6, WHITE, ncols=NCOLS)
                fill_row(ws, 7, WHITE, ncols=NCOLS)

                now = datetime.now()
                mset(ws, 6, 2, 6, 2, "📅  Report Period :",
                     bold=False, size=10, color=GRAY_MID, bg=WHITE, halign="right")
                mset(ws, 6, 3, 6, 5, now.strftime("%B %Y"),
                     bold=True, size=12, color=RED_MAIN, bg=WHITE, halign="left")

                mset(ws, 7, 2, 7, 2, "🖨  Generated on :",
                     bold=False, size=10, color=GRAY_MID, bg=WHITE, halign="right")
                mset(ws, 7, 3, 7, 5, now.strftime("%d %B %Y  |  %H:%M WIB"),
                     bold=False, size=10, color=GRAY_TEXT, bg=WHITE, halign="left")

                fill_row(ws, 8, WHITE, ncols=NCOLS)

                # CONTENTS heading
                fill_row(ws, 9, WHITE, ncols=NCOLS)
                c9 = ws.cell(row=9, column=2, value="📋  SHEET CONTENTS")
                c9.font = Font(name="Calibri", bold=True, size=12, color=RED_MAIN)
                c9.fill = _fill(WHITE)
                c9.alignment = _align("left")
                ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=5)
                for col in range(2, NCOLS):
                    ws.cell(row=9, column=col).border = Border(
                        bottom=Side(style="medium", color=RED_MAIN)
                    )

                fill_row(ws, 10, WHITE, ncols=NCOLS)

                # table header: No | Sheet Name | Rows | Cols
                fill_row(ws, 11, WHITE, ncols=NCOLS)
                for col, text in [(2, "No"), (3, "Sheet Name"), (4, "Total Rows"), (5, "Total Cols")]:
                    c = ws.cell(row=11, column=col, value=text)
                    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
                    c.fill = _fill(RED_MAIN)
                    c.alignment = _align("center")
                    c.border = _border(RED_MAIN)

                # data rows
                for i, (key, df) in enumerate(valid):
                    row = 12 + i
                    stripe = RED_LIGHT if i % 2 == 0 else WHITE
                    fill_row(ws, row, stripe, ncols=NCOLS)

                    c = ws.cell(row=row, column=2, value=i + 1)
                    c.font = Font(name="Calibri", bold=True, size=10, color=RED_MAIN)
                    c.fill = _fill(stripe); c.alignment = _align("center"); c.border = _border()

                    c = ws.cell(row=row, column=3, value=key)
                    c.font = Font(name="Calibri", size=10, color=GRAY_TEXT)
                    c.fill = _fill(stripe); c.alignment = _align("left"); c.border = _border()

                    c = ws.cell(row=row, column=4, value=len(df))
                    c.font = Font(name="Calibri", size=10, color=GRAY_TEXT)
                    c.fill = _fill(stripe); c.alignment = _align("center"); c.border = _border()

                    c = ws.cell(row=row, column=5, value=len(df.columns))
                    c.font = Font(name="Calibri", size=10, color=GRAY_TEXT)
                    c.fill = _fill(stripe); c.alignment = _align("center"); c.border = _border()

                # footer
                footer_row = 12 + len(valid)
                fill_row(ws, footer_row, RED_DARK, ncols=NCOLS)
                mset(ws, footer_row, 2, footer_row, 5,
                     "© INVENTRA  —  Confidential",
                     bold=False, size=9, color=RED_MID,
                     bg=RED_DARK, halign="center", italic=True)

            # ── DATA SHEET (ultra-fast: style header only, data raw) ──
            def write_styled_sheet(ws, df, title, now=None):
                if now is None:
                    now = datetime.now()
                ws.sheet_view.showGridLines = False
                nc = len(df.columns)
                nr = len(df)

                # ── title bar (row 1) ──
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
                t = ws.cell(row=1, column=1, value=f"  {title}")
                t.font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
                t.fill      = _fill(RED_DARK)
                t.alignment = _align("left")
                ws.row_dimensions[1].height = 28

                # ── subtitle (row 2) ──
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
                s = ws.cell(row=2, column=1)
                if title == "ANALISIS NON SETTING":
                    prev_str = (
                        f" ---------- previous result : {analisis_file_date.strftime('%d %B %Y, %H:%M')}"
                        if analisis_file_date else ""
                    )
                    s.value = (
                        f"  🖨  Generated on :   {now.strftime('%d %B %Y')}  |  {now.strftime('%H:%M')} WIB"
                        f"{prev_str}"
                    )
                else:
                    s.value = f"  Generated: {now.strftime('%d %B %Y, %H:%M')}"
                s.font      = Font(name="Calibri", size=9, italic=True, color=GRAY_MID)
                s.fill      = _fill(GRAY_LT)
                s.alignment = _align("left")
                ws.row_dimensions[2].height = 16
                ws.row_dimensions[3].height = 6   # spacer

                # ── header kolom (row 4) — hanya 1 baris, cepat ──
                for ci, col_name in enumerate(df.columns, 1):
                    c = ws.cell(row=4, column=ci, value=str(col_name))
                    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=11)
                    c.fill      = _fill(RED_MAIN)
                    c.alignment = _align("center")
                    c.border    = _border()
                ws.row_dimensions[4].height = 30

                # ── tulis data: value only, tanpa styling (tercepat) ──
                # konversi ke list of lists dulu (pandas vectorized, jauh lebih cepat dari iterrows)
                data_values = df.fillna("").values.tolist()
                for ri, row_vals in enumerate(data_values, 5):
                    for ci, value in enumerate(row_vals, 1):
                        ws.cell(row=ri, column=ci, value=value)

                # ── freeze & autofilter ──
                ws.freeze_panes = ws.cell(row=5, column=1)
                ws.auto_filter.ref = f"A4:{get_column_letter(nc)}{4 + nr}"

                # ── auto-fit: sampling 100 baris + header ──
                sample = df.head(100)
                for ci, col_name in enumerate(df.columns, 1):
                    header_len = len(str(col_name))
                    data_len   = sample.iloc[:, ci - 1].astype(str).str.len().max()
                    data_len   = 0 if pd.isna(data_len) else int(data_len)
                    best       = min(max(max(header_len, data_len) + 3, 10), 45)
                    ws.column_dimensions[get_column_letter(ci)].width = best

            # ── BUILD WORKBOOK ───────────────────────────────────
            wb = Workbook()
            wb.remove(wb.active)

            # Timestamp sekali, dipakai cover & semua sheet
            now_export = datetime.now()

            # Cover sheet pertama
            ws_cover = wb.create_sheet("🏠 COVER", 0)
            write_cover(ws_cover, data_store, now_export)

            # Data sheets
            used_sheet_names = {"🏠 COVER"}
            for key, df in data_store.items():
                if df is None or not isinstance(df, pd.DataFrame):
                    continue

                sheet_name = str(key)[:31]
                original_name = sheet_name
                counter = 1
                while sheet_name in used_sheet_names:
                    sheet_name = f"{original_name[:28]}_{counter}"
                    counter += 1
                used_sheet_names.add(sheet_name)

                ws = wb.create_sheet(sheet_name)
                write_styled_sheet(ws, df, key, now_export)

            wb.save(file_path)
            self.root.after(0, lambda: self._save_success(file_path))

        except Exception as e:
            err_msg = str(e)
            self.root.after(0, lambda: self._save_error(err_msg))

    # ==============================
    # SUCCESS / ERROR HANDLER
    # ==============================
    def _save_success(self, file_path):
        self.hide_processing_overlay()

        if hasattr(self, "btn_save"):
            self.btn_save.config(state="normal")

        messagebox.showinfo("Success", f"Data saved to:\n{file_path}")

    def _save_error(self, error_message):
        self.hide_processing_overlay()

        if hasattr(self, "btn_save"):
            self.btn_save.config(state="normal")

        messagebox.showerror("Error", error_message)

    # ==============================
    # OVERLAY
    # ==============================
    def show_processing_overlay(self, message="Processing data...\nPlease wait"):

        # Prevent duplicate overlay
        if hasattr(self, "processing_overlay") and self.processing_overlay.winfo_exists():
            return

        self.processing_overlay = tk.Toplevel(self.root)
        self.processing_overlay.overrideredirect(True)
        self.processing_overlay.attributes("-alpha", 0.85)
        self.processing_overlay.configure(bg="black")

        self.processing_overlay.geometry(
            f"{self.root.winfo_width()}x{self.root.winfo_height()}+"
            f"{self.root.winfo_rootx()}+{self.root.winfo_rooty()}"
        )

        label = tk.Label(
            self.processing_overlay,
            text=message,
            font=("Segoe UI", 16, "bold"),
            bg="black",
            fg="white"
        )
        label.pack(expand=True)

        self.processing_overlay.lift()
        self.processing_overlay.grab_set()

    def hide_processing_overlay(self):
        if hasattr(self, "processing_overlay"):
            if self.processing_overlay.winfo_exists():
                self.processing_overlay.grab_release()
                self.processing_overlay.destroy()
    
    def reset_all(self):
        """Clear all data"""
        if messagebox.askyesno("Confirm", "Clear all uploaded files and data?"):
            self.uploaded_files.clear()
            
            for key in self.data_store:
                self.data_store[key] = None
            
            for file_id, card_info in list(self.file_cards.items()):
                card_info['frame'].destroy()
            self.file_cards.clear()
            self.file_counter = 0
            
            # Save to JSON
            self.save_to_json()
            
            self.show_empty_state()
            self.update_file_count()
            
            self.tree.delete(*self.tree.get_children())
            self.current_view = None
            
            for btn in self.view_buttons.values():
                btn.config(bg=self.colors['border'], fg=self.colors['text'])
            
            self.info_label.config(text="All data cleared")
    
    def read_excel_auto_header(self, file_path, sheet_name):
        """
        Baca excel dan otomatis cari baris header.
        Baris dianggap header jika mengandung 'Stock Code' DAN
        salah satu variasi kata 'District' (Dstrct, Distric, District, dll).
        """

        # baca tanpa header dulu
        temp_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        header_row = None

        # Pola district: cocok dengan variasi seperti Dstrct, Distric, District, Distrct
        district_pattern = r"dis?tri?c"

        for i, row in temp_df.iterrows():
            row_str = row.astype(str)
            has_stock_code = row_str.str.contains(r"stock\s*code", case=False, na=False, regex=True).any()
            has_district   = row_str.str.contains(district_pattern, case=False, na=False, regex=True).any()

            if has_stock_code and has_district:
                header_row = i
                break

        if header_row is None:
            # fallback normal
            return pd.read_excel(file_path, sheet_name=sheet_name)

        # baca ulang dengan header ditemukan
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)

        # 🔥 NORMALISASI HEADER DI SINI
        df.columns = (
            df.columns
            .str.strip()
            .str.replace("\u00A0", " ", regex=False)  # ganti non-breaking space
        )

        return df



if __name__ == "__main__":
    root = tk.Tk()
    app = ModernRedINVENTRAManager(root)
    root.mainloop()